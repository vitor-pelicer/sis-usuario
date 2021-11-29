
from django.shortcuts import render, redirect
from conta.models import Usuario
import secrets
from django.contrib.auth import authenticate, login
from django.contrib import messages

#download xslx
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from django.http import HttpResponse

def login_view(request):
    autenticador=False
    if request.method == "POST":
        user = request.POST.get("user")
        senha = request.POST.get("password")
        try:
            u1 = Usuario.objects.get(username=user)
            p1 = u1.password
            if p1 == senha:
                autenticador = True
            else:
                autenticador=False
        except:
            autenticador = False
        if autenticador:
            messages.success(request, 'Logado com sucesso')
        else:
            context = {"error" : "Usuario e/ou senha invalidos"}
            return render(request, "conta/tela_de_login.html", context)
    return render(request, "conta/tela_de_login.html", {})

def cadastro_view(request):
    if request.method == "POST":
        nome = request.POST.get("nome")
        user = request.POST.get("user")
        data_nasc = request.POST.get("data_nasc")
        senha = request.POST.get("password")
        if(nome=='' or user=='' or data_nasc==''):
            context = {"error" : "Campos obrigatórios estão em branco"}
            return render(request, "conta/tela_de_cadastro.html", context)
        try:
            Usuario.objects.get(username=user)
            context = {"error" : "Nome de usuario já existente"}
            return render(request, "conta/tela_de_cadastro.html", context)
        except:
            if senha == '':
                senha = secrets.token_urlsafe(10)
                messages.success(request, 'Cadastrado com sucesso. Sua senha é : ' + senha)
            else:
                messages.success(request, 'Cadastrado com sucesso')
            new_user = Usuario.objects.create(nome=nome, username=user, nascimento=data_nasc, password=senha)
        
    return render(request, "conta/tela_de_cadastro.html", {})

def download_view(request):
    if request.method == "POST":
        return download()
    return render(request, "conta/tela_download.html")
        
def home_view(request):
    return render(request, "conta/tela_home.html")


def download():
    user_queryset = Usuario.objects.all()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Usuarios-{date}.xlsx'.format(
        date=datetime.now().strftime('%d-%m-%Y'),
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Usuarios'
    columns = [
        'Nome',
        'Usuario',
        'Data de nascimento',
        'Senha',
    ]

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for user in user_queryset:
        row_num+=1
        row=[
            user.nome,
            user.username,
            user.nascimento,
            user.password,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

